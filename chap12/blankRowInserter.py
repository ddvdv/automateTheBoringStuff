#! /usr/bin/pyton3
# description

import openpyxl, sys

# open file passed in arg
pathToUpdate = './' + str(sys.argv[3])
insertFromRow = int(sys.argv[1])
numRows = int(sys.argv[2])

wb = openpyxl.load_workbook(pathToUpdate)
ws = wb.active

updatedWb = openpyxl.Workbook()
updatedWb.create_sheet()
updatedWs = updatedWb.active

for rowNum in range(1, ws.max_row):
	if rowNum < insertFromRow:
		for colNum in range(1, 100):
			updatedWs.cell(row=rowNum, column=colNum).value = ws.cell(row=rowNum, column=colNum).value
	else:
		for colNum in range(1, 100):
			updatedWs.cell(row=rowNum + numRows , column=colNum).value = ws.cell(row=rowNum, column=colNum).value

updatedWb.save('updatedAfterInsertion.xlsx')
