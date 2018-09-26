#! /usr/bin/pyton3

import openpyxl, sys, re

# open file passed in arg
fileToUpdate = str(sys.argv[1])

def inverted(targetFile):
	wb = openpyxl.load_workbook(targetFile)
	ws = wb.active
	extractBaseReg = re.compile(r'(.*)\.(.*)')
	resultSearch = extractBaseReg.search(targetFile)
	resultBase = resultSearch.group(1)
	exportedText = open(resultBase + '.txt', 'w')

	for rowNum in range(1, ws.max_row):
		for colNum in range(1, ws.max_column):
			exportedText.write(str(ws.cell(row=rowNum, column=colNum).value) + ' --- ')
		exportedText.write('\n')

	exportedText.close()

inverted(fileToUpdate)
