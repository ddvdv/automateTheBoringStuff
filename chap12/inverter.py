#! /usr/bin/pyton3

import openpyxl, sys

# open file passed in arg
fileToUpdate = str(sys.argv[1])

def inverted(targetFile):
	wb = openpyxl.load_workbook(targetFile)
	ws = wb.active

	invertedWb = openpyxl.Workbook()
	invertedWb.create_sheet()
	invertedWs = invertedWb.active

	for rowNum in range(1, ws.max_row):
		for colNum in range(1, ws.max_column):
			invertedWs.cell(row=rowNum, column=colNum).value = ws.cell(row=colNum, column=rowNum).value

	invertedWb.save('invertedResult.xlsx')

inverted(fileToUpdate)
